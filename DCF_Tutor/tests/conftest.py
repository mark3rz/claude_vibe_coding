"""Pytest fixtures — synthetic CIQ-format workbook generator."""

from __future__ import annotations

import io
import pytest
import pandas as pd
import numpy as np
import openpyxl


def _build_synthetic_workbook() -> io.BytesIO:
    """Create a synthetic multi-sheet Excel workbook mimicking CIQ format."""
    buf = io.BytesIO()
    wb = openpyxl.Workbook()

    years = list(range(2019, 2025))
    year_headers = [f"{y} FY" for y in years]

    # --- Helper to write metadata rows ---
    def _write_meta(ws, company="Acme Corp", currency="USD", magnitude="Thousands (K)"):
        ws.append([""])  # row 1
        ws.append([company])  # row 2
        for _ in range(7):  # rows 3-9
            ws.append([""])
        ws.append([f"Currency: {currency}"])  # row 10
        ws.append([f"Magnitude: {magnitude}"])  # row 11
        ws.append([""])  # row 12

    # ═══════════════════════════════════════════════════════════════
    # Sheet 0: Income Statement
    # ═══════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Income Statement"
    _write_meta(ws)

    # Row 13: year headers
    ws.append([""] + year_headers)

    # Row 14: blank
    ws.append([""])

    # Financial rows
    data = {
        "Revenue": [8000, 9000, 10000, 11000, 12000, 13000],
        "Cost of Goods Sold": [-4000, -4500, -5000, -5300, -5700, -6100],
        "Gross Profit": [4000, 4500, 5000, 5700, 6300, 6900],
        "Selling, General and Administrative": [-1500, -1600, -1700, -1800, -1900, -2000],
        "Research and Development": [-500, -550, -600, -650, -700, -750],
        "EBITDA": [2400, 2700, 3000, 3600, 4100, 4550],
        "Depreciation and Amortization": [-400, -420, -450, -480, -520, -560],
        "EBIT": [2000, 2280, 2550, 3120, 3580, 3990],
        "Pretax Income": [1800, 2100, 2350, 2900, 3300, 3700],
        "Income Tax Expense": [-378, -441, -494, -609, -693, -777],
        "Net Income": [1422, 1659, 1856, 2291, 2607, 2923],
        "Stock Based Compensation": [100, 120, 140, 160, 180, 200],
        "EPS - Diluted": [2.84, 3.32, 3.71, 4.58, 5.21, 5.85],
        "Diluted Shares Outstanding": [500000000, 500000000, 500000000, 500000000, 500000000, 500000000],
    }
    for label, vals in data.items():
        ws.append([label] + vals)

    # ═══════════════════════════════════════════════════════════════
    # Sheet 1: Balance Sheet
    # ═══════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Balance Sheet")
    _write_meta(ws2)
    ws2.append([""] + year_headers)
    ws2.append([""])

    bs_data = {
        "Cash and Equivalents": [3000, 3500, 4200, 5000, 5800, 6500],
        "Total Current Assets": [5000, 5500, 6200, 7000, 7800, 8500],
        "Net Property, Plant and Equipment": [4000, 4200, 4500, 4900, 5300, 5700],
        "Total Assets": [15000, 16000, 17500, 19000, 20500, 22000],
        "Total Current Liabilities": [3000, 3200, 3500, 3800, 4100, 4400],
        "Total Debt": [5000, 4800, 4500, 4200, 3900, 3600],
        "Net Debt": [2000, 1300, 300, -800, -1900, -2900],
        "Total Liabilities": [9000, 9200, 9400, 9600, 9700, 9800],
        "Total Stockholders Equity": [6000, 6800, 8100, 9400, 10800, 12200],
    }
    for label, vals in bs_data.items():
        ws2.append([label] + vals)

    # ═══════════════════════════════════════════════════════════════
    # Sheet 2: Cash Flow
    # ═══════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Cash Flow Statement")
    _write_meta(ws3)
    ws3.append([""] + year_headers)
    ws3.append([""])

    cf_data = {
        "Cash from Operating Activities": [2200, 2500, 2800, 3200, 3600, 4000],
        "Capital Expenditures": [-500, -550, -600, -650, -700, -750],
        "Levered Free Cash Flow": [1700, 1950, 2200, 2550, 2900, 3250],
        "Change in Net Working Capital": [-100, -120, -140, -160, -180, -200],
        "Depreciation and Amortization": [400, 420, 450, 480, 520, 560],
    }
    for label, vals in cf_data.items():
        ws3.append([label] + vals)

    # ═══════════════════════════════════════════════════════════════
    # Sheet 3: Multiples
    # ═══════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("Valuation Multiples")
    _write_meta(ws4)
    mult_years = [f"{y} FY" for y in range(2022, 2025)]
    ws4.append([""] + mult_years)
    ws4.append([""])

    mult_data = {
        "TEV/Total Revenue": [3.5, 3.2, 3.0],
        "TEV/EBITDA": [12.5, 11.0, 10.5],
        "P/E": [20.0, 18.5, 17.0],
        "Price/Book": [4.0, 3.5, 3.2],
        "Market Capitalization": [28000, 26000, 25000],
        "Total Enterprise Value": [30000, 27300, 25300],
    }
    for label, vals in mult_data.items():
        ws4.append([label] + vals)

    # ═══════════════════════════════════════════════════════════════
    # Sheet 4: Performance Analysis
    # ═══════════════════════════════════════════════════════════════
    ws5 = wb.create_sheet("Performance Analysis")
    _write_meta(ws5)
    ws5.append([""] + year_headers)
    ws5.append([""])
    ws5.append(["Gross Margin"] + [50.0, 50.0, 50.0, 51.8, 52.5, 53.1])
    ws5.append(["EBITDA Margin"] + [30.0, 30.0, 30.0, 32.7, 34.2, 35.0])

    wb.save(buf)
    buf.seek(0)
    return buf


@pytest.fixture
def synthetic_xlsx() -> io.BytesIO:
    """Return a BytesIO with a synthetic CIQ Excel workbook."""
    return _build_synthetic_workbook()


@pytest.fixture
def parsed_workbook(synthetic_xlsx):
    """Return a parsed RawWorkbook from the synthetic fixture."""
    from src.importer import parse_workbook
    return parse_workbook(synthetic_xlsx)


@pytest.fixture
def standardized_data(parsed_workbook):
    """Return (StandardizedFinancials, MappingReport, exit_multiple_suggestion)."""
    from src.mapping import build_standardized
    return build_standardized(parsed_workbook, include_estimates=False, threshold=0.70)
